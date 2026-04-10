"""
재무 라우터 — /api/finance/*
"""
import json
import uuid
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from services.finance.finance_ocr_service import analyze_receipt
from database import get_connection

router = APIRouter()

# 이미지 저장 경로
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)


# ──────────────────────────────────────────────────────────────
# 요청 스키마
# ──────────────────────────────────────────────────────────────
class TransactionItem(BaseModel):
    item: str
    amount: int
    tax_amount: int
    account_code: str = "기타비용"
    memo: str = ""
    confidence: float = 0.0


class SaveTransactionsRequest(BaseModel):
    receipt_date: Optional[str] = None
    vendor: str = ""
    image_path: Optional[str] = None   # OCR 저장 시 반환된 서버 경로
    items: list[TransactionItem]


class UpdateTransactionRequest(BaseModel):
    account_code: Optional[str] = None
    amount: Optional[int] = None
    tax_amount: Optional[int] = None
    memo: Optional[str] = None


# ──────────────────────────────────────────────────────────────
# POST /api/finance/ocr  — 이미지 저장 + OCR 분석 + 중복 탐지
# ──────────────────────────────────────────────────────────────
@router.post("/ocr")
async def ocr(file: UploadFile = File(...)):
    """
    영수증 파일을 저장하고 AI로 분석합니다. DB 저장은 하지 않습니다.
    중복 의심 항목은 is_duplicate: true로 표시되어 반환됩니다.
    """
    file_bytes = await file.read()
    content_type = file.content_type or "image/jpeg"

    if not file_bytes:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    # 이미지를 uploads/ 에 저장
    ext = Path(file.filename or "receipt").suffix or ".jpg"
    filename = f"{date.today()}_{uuid.uuid4().hex[:8]}{ext}"
    save_path = UPLOAD_DIR / filename
    save_path.write_bytes(file_bytes)
    image_path = f"uploads/{filename}"

    # OCR 분석
    try:
        result = analyze_receipt(file_bytes, content_type)
    except Exception as e:
        save_path.unlink(missing_ok=True)   # 분석 실패 시 이미지 삭제
        raise HTTPException(status_code=502, detail=f"AI 분석 실패: {str(e)}")

    receipt_date = result.get("receipt_date") or str(date.today())
    vendor       = result.get("vendor", "")
    items_raw    = result.get("items", [])

    if not items_raw:
        save_path.unlink(missing_ok=True)
        raise HTTPException(status_code=422, detail="영수증에서 항목을 인식하지 못했습니다.")

    # 중복 탐지 — vendor + receipt_date + amount 동시 일치 여부 조회
    dup_amounts: set[int] = set()
    if vendor and receipt_date:
        amounts = [int(it.get("amount", 0)) for it in items_raw]
        if amounts:
            conn = get_connection()
            cur  = conn.cursor()
            try:
                placeholders = ", ".join(["%s"] * len(amounts))
                cur.execute(
                    f"""
                    SELECT DISTINCT amount FROM finance_transactions
                    WHERE vendor = %s AND receipt_date = %s
                      AND amount IN ({placeholders})
                    """,
                    [vendor, receipt_date] + amounts,
                )
                dup_amounts = {r[0] for r in cur.fetchall()}
            finally:
                cur.close()
                conn.close()

    items = [
        {
            "item":         it.get("item", ""),
            "amount":       int(it.get("amount", 0)),
            "tax_amount":   int(it.get("tax_amount", 0)),
            "total_amount": int(it.get("amount", 0)) + int(it.get("tax_amount", 0)),
            "account_code": it.get("account_code", "기타비용"),
            "vendor":       vendor,
            "memo":         it.get("memo", ""),
            "confidence":   float(it.get("confidence", 0)),
            "is_duplicate": int(it.get("amount", 0)) in dup_amounts,
        }
        for it in items_raw
    ]

    return {
        "receipt_date":  receipt_date,
        "vendor":        vendor,
        "image_path":    image_path,
        "items":         items,
        "has_duplicates": len(dup_amounts) > 0,
    }


# ──────────────────────────────────────────────────────────────
# POST /api/finance/transactions  — 분석 결과를 DB에 저장
# ──────────────────────────────────────────────────────────────
@router.post("/transactions", status_code=201)
def save_transactions(body: SaveTransactionsRequest):
    """
    OCR 분석 결과를 DB에 저장합니다. status 기본값은 'pending'.
    """
    if not body.items:
        raise HTTPException(status_code=400, detail="저장할 항목이 없습니다.")

    receipt_date = body.receipt_date or str(date.today())
    vendor       = body.vendor
    saved        = []

    conn = get_connection()
    cur  = conn.cursor()

    try:
        for item in body.items:
            cur.execute(
                """
                INSERT INTO finance_transactions
                    (receipt_date, item, amount, tax_amount,
                     account_code, vendor, memo, ai_confidence, raw_json, image_path)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, total_amount, created_at
                """,
                (
                    receipt_date,
                    item.item,
                    item.amount,
                    item.tax_amount,
                    item.account_code,
                    vendor,
                    item.memo,
                    item.confidence,
                    json.dumps(item.model_dump(), ensure_ascii=False),
                    body.image_path,
                ),
            )
            row = cur.fetchone()
            saved.append({
                "id":           row[0],
                "receipt_date": receipt_date,
                "item":         item.item,
                "amount":       item.amount,
                "tax_amount":   item.tax_amount,
                "total_amount": row[1],
                "account_code": item.account_code,
                "vendor":       vendor,
                "memo":         item.memo,
                "confidence":   item.confidence,
                "image_path":   body.image_path,
                "status":       "pending",
                "created_at":   str(row[2]),
            })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 저장 실패: {str(e)}")

    finally:
        cur.close()
        conn.close()

    return {"saved": saved}


# ──────────────────────────────────────────────────────────────
# GET /api/finance/transactions/export  — 확정 전표 엑셀 다운로드
# (주의: /{id} 라우트보다 먼저 선언해야 라우팅 충돌 없음)
# ──────────────────────────────────────────────────────────────
@router.get("/transactions/export")
def export_confirmed_excel():
    """
    status = 'confirmed' 인 전표만 엑셀(.xlsx)로 다운로드합니다.
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            SELECT id, receipt_date, item, amount, tax_amount, total_amount,
                   account_code, vendor, memo, ai_confidence, created_at
            FROM finance_transactions
            WHERE status = 'confirmed'
            ORDER BY receipt_date DESC, id DESC
            """,
        )
        rows = cur.fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"조회 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "확정 전표"

    # 헤더 스타일
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["ID", "날짜", "항목", "공급가액", "부가세", "합계", "계정과목", "거래처", "적요", "신뢰도", "등록일"]
    col_widths = [6, 12, 30, 12, 10, 12, 14, 20, 25, 8, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    # 데이터 행
    for row in rows:
        ws.append([
            row[0],          # id
            str(row[1]),     # receipt_date
            row[2],          # item
            row[3],          # amount
            row[4],          # tax_amount
            row[5],          # total_amount
            row[6],          # account_code
            row[7] or "",    # vendor
            row[8] or "",    # memo
            float(row[9]) if row[9] else None,  # ai_confidence
            str(row[10]),    # created_at
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''confirmed_transactions.xlsx"},
    )


# ──────────────────────────────────────────────────────────────
# GET /api/finance/transactions  — DB 전표 내역 조회
# ──────────────────────────────────────────────────────────────
@router.get("/transactions")
def list_transactions(
    limit:        int            = Query(default=50, le=200),
    offset:       int            = Query(default=0, ge=0),
    account_code: Optional[str]  = Query(default=None),
    status:       Optional[str]  = Query(default=None),
    date_from:    Optional[str]  = Query(default=None),
    date_to:      Optional[str]  = Query(default=None),
):
    """
    DB에 저장된 전표 목록을 반환합니다.

    Query params:
      limit        (int, default 50)
      offset       (int, default 0)
      account_code (str, optional)
      status       (str, optional) — 'pending' | 'confirmed'
      date_from    (str YYYY-MM-DD, optional)
      date_to      (str YYYY-MM-DD, optional)
    """
    where_clauses = []
    params = []

    if account_code:
        where_clauses.append("account_code = %s")
        params.append(account_code)
    if status:
        where_clauses.append("status = %s")
        params.append(status)
    if date_from:
        where_clauses.append("receipt_date >= %s")
        params.append(date_from)
    if date_to:
        where_clauses.append("receipt_date <= %s")
        params.append(date_to)

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(f"SELECT COUNT(*) FROM finance_transactions {where_sql}", params)
        total = cur.fetchone()[0]

        cur.execute(
            f"""
            SELECT id, receipt_date, item, amount, tax_amount, total_amount,
                   account_code, vendor, memo, ai_confidence, status, image_path, created_at
            FROM finance_transactions
            {where_sql}
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
            """,
            params + [limit, offset],
        )
        rows = cur.fetchall()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"조회 실패: {str(e)}")

    finally:
        cur.close()
        conn.close()

    return {
        "total": total,
        "items": [
            {
                "id":           r[0],
                "receipt_date": str(r[1]),
                "item":         r[2],
                "amount":       r[3],
                "tax_amount":   r[4],
                "total_amount": r[5],
                "account_code": r[6],
                "vendor":       r[7],
                "memo":         r[8],
                "confidence":   float(r[9]) if r[9] else None,
                "status":       r[10],
                "image_path":   r[11],
                "created_at":   str(r[12]),
            }
            for r in rows
        ],
    }


# ──────────────────────────────────────────────────────────────
# PUT /api/finance/transactions/{id}  — 전표 수정
# ──────────────────────────────────────────────────────────────
@router.put("/transactions/{transaction_id}")
def update_transaction(transaction_id: int, body: UpdateTransactionRequest):
    """
    계정과목, 금액, 부가세, 적요를 수정합니다.
    """
    set_clauses = []
    params      = []

    if body.account_code is not None:
        set_clauses.append("account_code = %s"); params.append(body.account_code)
    if body.amount is not None:
        set_clauses.append("amount = %s");       params.append(body.amount)
    if body.tax_amount is not None:
        set_clauses.append("tax_amount = %s");   params.append(body.tax_amount)
    if body.memo is not None:
        set_clauses.append("memo = %s");         params.append(body.memo)

    if not set_clauses:
        raise HTTPException(status_code=400, detail="수정할 항목이 없습니다.")

    params.append(transaction_id)
    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            f"""
            UPDATE finance_transactions
               SET {', '.join(set_clauses)}, updated_at = NOW()
             WHERE id = %s
            RETURNING id, account_code, amount, tax_amount, total_amount, memo, updated_at
            """,
            params,
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="전표를 찾을 수 없습니다.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"수정 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {
        "id":           row[0],
        "account_code": row[1],
        "amount":       row[2],
        "tax_amount":   row[3],
        "total_amount": row[4],
        "memo":         row[5],
        "updated_at":   str(row[6]),
    }


# ──────────────────────────────────────────────────────────────
# POST /api/finance/transactions/{id}/confirm  — 최종 확정
# ──────────────────────────────────────────────────────────────
@router.post("/transactions/{transaction_id}/confirm")
def confirm_transaction(transaction_id: int):
    """
    전표 status를 'confirmed'로 업데이트합니다.
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            "UPDATE finance_transactions SET status = 'confirmed', updated_at = NOW() WHERE id = %s RETURNING id, status",
            (transaction_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="전표를 찾을 수 없습니다.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"확정 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {"id": row[0], "status": row[1]}
